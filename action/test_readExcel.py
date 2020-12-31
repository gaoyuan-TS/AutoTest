#!/usr/bin/env python
# encoding: utf-8
"""
@author :gaoyuan
@contact:1103313679@qq.com
@time   :2020/11/11 10:12
"""
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import xlrd,xlwt
import time
from openpyxl import load_workbook

from action.ObjectMap import ObjectMap


class Option():
    def openBrower(self,brower):
        if brower == 'chrome':
            path = r'D:\AutoTest\drivers\chrome\86.0.4240.22\chromedriver.exe'
            option = Options()
            self.driver = webdriver.Chrome(executable_path=path,options=option)
        else:
            path = r'D:\AutoTest\drivers\firefox\geckodriver.exe'
            option = Options()
            self.driver = webdriver.Firefox(executable_path=path,options=option)



    def input(self,locator,value):
        self.driver.find_element_by_xpath(locator).send_keys(value)


    def click(self,locator):
        self.driver.find_element_by_xpath(locator).click()



class parselExcel(object):
    def __init__(self,file_path):
        self.file_path = file_path
        self.wb = load_workbook(file_path)

    def write_cell_value(self, sheetname,  rowno,  columnno, value):
        try:
            self.wb_sheet = self.wb[sheetname]
            self.wb_sheet.cell(rowno,  columnno, value)
            self.wb.save(self.file_path)
        except Exception as e:
            print(e)



    def ismerge(self, sheetname):
        """
        判断'工作表'内是否有合并单元格
        :param sheetname:
        :return:
        """
        sheetnames = self.wb[sheetname]
        merge = sheetnames.merged_cells
        for x in merge:
            print(x)
        return merge

    def read_excel(self,file_path):
        #r'D:\AutoTest\exceltemplate\test.xlsx'
        book = xlrd.open_workbook(file_path)
        # book_sheet = book.sheet_by_name('')
        book_sheet = book.sheet_by_index(0)
        nrow = book_sheet.nrows
        ncol =len(book_sheet.row_values(0))
        accout = {}
        print(nrow,ncol)
        list =[]
        for x in range(1,nrow ):
            row_value = book_sheet.row_values(x)
            row_dir = []
            for y in range(0,ncol):
                cell_value = book_sheet.cell(x,y)
                row_dir.append(cell_value.value)
            list.append(row_dir)
        print(list[0:nrow])
        # print(list[0])


    def getColumnValue(self,sheetname,column):
        """

        :return: 获取一列的值
        """
        sheetnames = self.wb[sheetname]
        columns = sheetnames.max_row+1
        list = []
        for x in range(1,columns+1):
            value1 = sheetnames.cell(x,column).value
            list.append(value1)

        for index,value in enumerate(list,1):
            print(index,value)
            if value == 'email':
                print('这是email')
                continue
            elif value is None or value == ''or value == 'None':
                self.write_cell_value(sheetname,index,column,'往空白填入数值')
        return list


p = parselExcel(r'D:\AutoTest\exceltemplate\test.xlsx')
# p.write_cell_value('Sheet1',5,5,'测试写入数据')
list = p.ismerge('Sheet1')
print(list)
# sheetnames = p.wb.sheetnames
# p.getColumnValue('Sheet1',4)

# p = Option()
# p.openBrower('chrome')
# p.driver.get('http://zs-beta.cntracechain.com/#/login')
# p.driver.maximize_window()
# try:
#     # p.driver.implicity_wait(10)
#     p.input('/html/body/div/div/section/main/div/div[2]/div/div/div[3]/form/div/div[2]/div/div/input','gaoyuan')
#     p.input('/html/body/div/div/section/main/div/div[2]/div/div/div[3]/form/div/div[3]/div/div/input','zs666666')
#     p.click('//*[@id="login"]/section/main/div/div[2]/div/div/div[3]/form/div/div[5]/div/button')
#     time.sleep(2)
# except Exception as e:
#     print(e)
# finally:
#     p.driver.quit()